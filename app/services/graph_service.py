# app/services/graph_service.py

import os
import requests
import openpyxl
import io
import re

# --- Конфигурация ---
# Эти переменные должны быть установлены в вашем .env файле для аутентификации
MS_CLIENT_ID = os.environ.get("MS_CLIENT_ID")
MS_CLIENT_SECRET = os.environ.get("MS_CLIENT_SECRET")
MS_TENANT_ID = os.environ.get("MS_TENANT_ID")

# Email или ID пользователя, чей OneDrive будет использоваться.
# Требуется для потока "client credentials" (доступ от имени приложения).
MS_ONEDRIVE_USER_ID = os.environ.get("MS_ONEDRIVE_USER_ID")


class GraphAPIError(Exception):
    """Пользовательское исключение для ошибок при работе с Graph API."""
    pass


def _get_access_token():
    """
    Выполняет аутентификацию в Microsoft Identity Platform для получения токена доступа.
    Использует поток "client credentials" (учетные данные клиента).
    """
    if not all([MS_CLIENT_ID, MS_CLIENT_SECRET, MS_TENANT_ID]):
        raise GraphAPIError(
            "В файле .env отсутствуют учетные данные Microsoft: "
            "MS_CLIENT_ID, MS_CLIENT_SECRET, MS_TENANT_ID."
        )

    url = f"https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token"
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {
        'client_id': MS_CLIENT_ID,
        'scope': 'https://graph.microsoft.com/.default',
        'client_secret': MS_CLIENT_SECRET,
        'grant_type': 'client_credentials'
    }

    try:
        response = requests.post(url, headers=headers, data=payload)
        response.raise_for_status()  # Вызовет исключение для кодов 4xx/5xx
    except requests.exceptions.RequestException as e:
        raise GraphAPIError(f"Ошибка сети при получении токена доступа: {e}")

    token_data = response.json()
    access_token = token_data.get('access_token')

    if not access_token:
        error_details = token_data.get('error_description', 'Нет дополнительной информации.')
        raise GraphAPIError(f"Не удалось получить токен доступа. Ответ сервера: {error_details}")

    return access_token


def download_file_from_onedrive(file_path_in_onedrive: str) -> bytes:
    """
    Скачивает файл из корневой папки OneDrive указанного пользователя.

    :param file_path_in_onedrive: Путь к файлу от корневой папки,
                                  например, '/Documents/Отчеты/data.xlsx'
    :return: Содержимое файла в виде байтов.
    """
    if not MS_ONEDRIVE_USER_ID:
        raise GraphAPIError("В файле .env отсутствует ID пользователя OneDrive (MS_ONEDRIVE_USER_ID).")

    access_token = _get_access_token()

    # Формат API для доступа к файлу в диске конкретного пользователя.
    # Требует прав уровня приложения, таких как Files.Read.All.
    # Двоеточие в пути обязательно для API.
    api_url = (
        f"https://graph.microsoft.com/v1.0/users/{MS_ONEDRIVE_USER_ID}/drive/root:"
        f"{file_path_in_onedrive}:/content"
    )

    headers = {'Authorization': f'Bearer {access_token}'}

    try:
        response = requests.get(api_url, headers=headers)
        
        if response.status_code == 404:
            raise FileNotFoundError(f"Файл не найден в OneDrive по пути: {file_path_in_onedrive}")
        
        response.raise_for_status() # Проверка на другие ошибки HTTP
        
        return response.content

    except requests.exceptions.RequestException as e:
        raise GraphAPIError(f"Ошибка сети при скачивании файла: {e}")


def read_row_from_excel_bytes(excel_bytes: bytes, row_number: int) -> dict:
    """
    Читает указанную строку из Excel-файла, переданного в виде байтов,
    и возвращает словарь вида {заголовок: значение}.

    :param excel_bytes: Содержимое .xlsx файла.
    :param row_number: Номер строки для чтения (нумерация с 1).
    :return: Словарь, сопоставляющий заголовки столбцов со значениями ячеек.
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        sheet = workbook.active
    except Exception as e:
        raise ValueError(f"Не удалось прочитать содержимое Excel-файла. Ошибка: {e}")

    if not (2 <= row_number <= sheet.max_row):
        raise IndexError(f"Номер строки {row_number} находится вне допустимого диапазона (от 2 до {sheet.max_row}).")

    # Читаем и очищаем заголовки из первой строки
    headers = []
    for cell in sheet[1]:
        if cell.value is not None:
            header_text = str(cell.value).strip()
            header_text = re.sub(r'\s+', ' ', header_text)  # Заменяем множественные пробелы на один
            headers.append(header_text)
    
    if not headers:
        raise ValueError("Не удалось прочитать заголовки из первой строки Excel-файла.")

    # Читаем значения из указанной строки
    row_values = [cell.value if cell.value is not None else "" for cell in sheet[row_number]]

    # Создаем словарь для подстановки в шаблон Word
    placeholders = {f"{{{{{headers[i]}}}}}": str(row_values[i]) for i in range(len(headers))}
    
    return placeholders